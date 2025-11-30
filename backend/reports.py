"""
Report generation: CSV, styled Excel, and Sitemap XML exports.
"""
import csv
import io
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils import get_column_letter


def sanitize_cell(value):
    """Prevent CSV/Excel formula injection (Zero-day mitigation)."""
    if isinstance(value, str) and value and value[0] in ['=', '+', '-', '@']:
        return f"'{value}"
    return value

def generate_csv(results: list) -> str:
    """Generate a complete CSV report from crawl results."""
    if not results:
        return ""
    
    # Flatten — remove nested objects
    flat_keys = [k for k in results[0].keys() if k not in ('internal_links', 'external_links', 'images', 'redirect_chain')]
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=flat_keys, extrasaction='ignore')
    writer.writeheader()
    for row in results:
        sanitized_row = {k: sanitize_cell(row.get(k, '')) for k in flat_keys}
        writer.writerow(sanitized_row)
    return output.getvalue()


def generate_issues_csv(results: list) -> str:
    """Generate a CSV containing only pages with SEO issues."""
    issues = []
    for r in results:
        page_issues = []
        if r.get('status_code', 0) != 200:
            page_issues.append(f"Status {r.get('status_code')}")
        if r.get('title_length', 0) == 0:
            page_issues.append("Missing title")
        if r.get('meta_desc_length', 0) == 0:
            page_issues.append("Missing meta description")
        if r.get('h1_count', 0) == 0:
            page_issues.append("Missing H1")
        if r.get('h1_count', 0) > 1:
            page_issues.append(f"Multiple H1 ({r.get('h1_count')})")
        if r.get('images_without_alt', 0) > 0:
            page_issues.append(f"{r.get('images_without_alt')} images missing alt")
        if r.get('redirect_count', 0) > 0:
            page_issues.append(f"{r.get('redirect_count')} redirects")
        if not r.get('canonical_url'):
            page_issues.append("Missing canonical")
        if r.get('response_time', 0) > 3.0:
            page_issues.append(f"Slow ({r.get('response_time', 0):.1f}s)")
        if r.get('is_akamai_bypass'):
            page_issues.append("Required Akamai bypass")
        
        if page_issues:
            issues.append({
                'url': sanitize_cell(r.get('url', '')),
                'status_code': r.get('status_code', 0),
                'issues': sanitize_cell(' | '.join(page_issues)),
                'issue_count': len(page_issues),
                'title': sanitize_cell(r.get('title', '')),
                'indexability': sanitize_cell(r.get('indexability', '')),
            })
    
    if not issues:
        return "url,message\n,No issues found!"
    
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=['url', 'status_code', 'issues', 'issue_count', 'title', 'indexability'])
    writer.writeheader()
    writer.writerows(issues)
    return output.getvalue()


def generate_excel(results: list, filename_base: str = "ultimatecrawler") -> bytes:
    """Generate a professionally styled Excel workbook from crawl results."""
    wb = openpyxl.Workbook()
    
    # ─── Styles ───
    hdr_fill = PatternFill('solid', fgColor='0A1628')
    hdr_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    data_font = Font(size=10, name='Calibri')
    good_fill = PatternFill('solid', fgColor='D4EDDA')
    warn_fill = PatternFill('solid', fgColor='FFF3CD')
    bad_fill = PatternFill('solid', fgColor='F8D7DA')
    bypass_fill = PatternFill('solid', fgColor='E8DAEF')
    alt_fill = PatternFill('solid', fgColor='F0F4F8')
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0'),
    )
    
    def style_header(ws, headers):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[1].height = 32
        ws.freeze_panes = 'A2'
    
    # ─── Sheet 1: All URLs ───
    ws = wb.active
    ws.title = 'All URLs'
    headers = ['URL', 'Status', 'Indexability', 'Title', 'Title Len', 'Meta Desc Len',
               'H1 Count', 'Word Count', 'Response Time', 'Internal Links', 'External Links',
               'Images', 'Missing Alt', 'Schema', 'Canonical', 'Akamai Bypass', 'Error']
    style_header(ws, headers)
    
    for ri, r in enumerate(results, 2):
        vals = [
            r.get('url', ''), r.get('status_code', 0), r.get('indexability', ''),
            r.get('title', ''), r.get('title_length', 0), r.get('meta_desc_length', 0),
            r.get('h1_count', 0), r.get('word_count', 0),
            round(r.get('response_time', 0), 2),
            r.get('internal_links_count', 0), r.get('external_links_count', 0),
            r.get('image_count', 0), r.get('images_without_alt', 0),
            r.get('schema_types', ''), r.get('canonical_url', ''),
            'Yes' if r.get('is_akamai_bypass') else 'No', r.get('error', '')
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=sanitize_cell(v))
            cell.font = data_font
            cell.border = thin_border
            if ri % 2 == 0:
                cell.fill = alt_fill
            # Conditional highlights
            if ci == 2:  # Status
                if v == 200: cell.fill = good_fill
                elif v >= 400: cell.fill = bad_fill
                elif v >= 300: cell.fill = warn_fill
            if ci == 16 and v == 'Yes':  # Akamai bypass
                cell.fill = bypass_fill
    
    widths = [55, 10, 15, 45, 10, 13, 10, 11, 13, 13, 13, 10, 12, 20, 45, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # ─── Sheet 2: Issues ───
    ws2 = wb.create_sheet('Issues')
    issue_headers = ['URL', 'Status', 'Issue', 'Severity']
    style_header(ws2, issue_headers)
    row_idx = 2
    for r in results:
        url = r.get('url', '')
        checks = [
            (r.get('status_code', 0) != 200, f"Status {r.get('status_code')}", 'Critical'),
            (r.get('title_length', 0) == 0, "Missing title", 'Critical'),
            (r.get('meta_desc_length', 0) == 0, "Missing meta description", 'Warning'),
            (r.get('h1_count', 0) == 0, "Missing H1", 'Warning'),
            (r.get('h1_count', 0) > 1, f"Multiple H1 ({r.get('h1_count')})", 'Warning'),
            (r.get('images_without_alt', 0) > 0, f"{r.get('images_without_alt')} images without alt", 'Minor'),
            (r.get('response_time', 0) > 3, f"Slow page ({r.get('response_time',0):.1f}s)", 'Warning'),
            (r.get('is_akamai_bypass', False), "Required Akamai CURL bypass", 'Info'),
        ]
        for cond, msg, sev in checks:
            if cond:
                ws2.cell(row=row_idx, column=1, value=url).font = data_font
                ws2.cell(row=row_idx, column=2, value=r.get('status_code', 0)).font = data_font
                ws2.cell(row=row_idx, column=3, value=msg).font = data_font
                sev_cell = ws2.cell(row=row_idx, column=4, value=sev)
                sev_cell.font = data_font
                sev_cell.fill = bad_fill if sev == 'Critical' else warn_fill if sev == 'Warning' else good_fill
                for ci in range(1, 5):
                    ws2.cell(row=row_idx, column=ci).border = thin_border
                row_idx += 1
    
    for i, w in enumerate([55, 10, 40, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    
    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
